"""
Envelope Reconciliation: Mar 2022 - Dec 2025
Broadridge P&M Postage and Material Recon
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# CONFIG
SRC = r"C:\Users\smendoza\Projects\Broadridge Envelopes\P&M Postage and Material Recon.xlsx"
OUT = r"C:\Users\smendoza\Projects\Broadridge Envelopes\Envelope Reconciliation Mar2022-Current.xlsx"
START = pd.Timestamp("2022-03-01")
END = pd.Timestamp("2025-12-01")

print("Loading source data...")
xls = pd.ExcelFile(SRC)

ep = pd.read_excel(xls, sheet_name="Envelopes Purchased")
ep["Month"] = pd.to_datetime(ep["Month"])
ep = ep[(ep["Month"] >= START) & (ep["Month"] <= END)].copy()

vd = pd.read_excel(xls, sheet_name="Volume Data")
vd["Month"] = pd.to_datetime(vd["Month"])
vd = vd[(vd["Month"] >= START) & (vd["Month"] <= END)].copy()

post = pd.read_excel(xls, sheet_name="Postage Data")
post["Month"] = pd.to_datetime(post["Month"])
post = post[(post["Month"] >= START) & (post["Month"] <= END)].copy()

print(f"  Envelopes Purchased rows (filtered): {len(ep):,}")
print(f"  Volume Data rows (filtered):         {len(vd):,}")
print(f"  Postage Data rows (filtered):        {len(post):,}")

# PART 1 - MONTHLY SUMMARY
print("Building monthly summary...")
purch_m = ep.groupby("Month")["Quantity Purchased"].sum().rename("Purchased")
vol_m = vd.groupby("Month")["Envelopes"].sum().rename("Volume Usage")
mail_m = post.groupby("Month")["Env_Mailed"].sum().rename("Postal Usage")
spoil_m = post.groupby("Month")["Spoils"].sum().rename("Spoilage")

full_idx = pd.date_range(START, END, freq="MS", name="Month")
monthly = pd.DataFrame(index=full_idx)
monthly = monthly.join(purch_m).join(vol_m).join(mail_m).join(spoil_m)
monthly = monthly.fillna(0).astype("int64")
monthly["Variance (Purch-Usage)"] = monthly["Purchased"] - monthly["Volume Usage"]
monthly["Cumulative Purchased"] = monthly["Purchased"].cumsum()
monthly["Cumulative Usage"] = monthly["Volume Usage"].cumsum()
monthly["Cumulative Variance"] = monthly["Variance (Purch-Usage)"].cumsum()
monthly_out = monthly.reset_index()
monthly_out["Month"] = monthly_out["Month"].dt.strftime("%b-%y")

# PART 2 - ANNUAL SUMMARY
print("Building annual summary...")
monthly_yr = monthly.copy()
monthly_yr["Year"] = monthly_yr.index.year
annual = monthly_yr.groupby("Year")[["Purchased", "Volume Usage", "Postal Usage", "Spoilage"]].sum()
annual["Variance"] = annual["Purchased"] - annual["Volume Usage"]
annual["Variance %"] = annual["Variance"] / annual["Purchased"]
annual = annual.reset_index()

# PART 3 - BY ENVELOPE TYPE
print("Building envelope type summary...")
by_type = ep.groupby("Envelope Description").agg(
    Total_Qty_Purchased=("Quantity Purchased", "sum"),
    Total_Receipt_Amount=("Receipt Amount", "sum"),
    Total_Invoiced=("Total Invoiced", "sum"),
).reset_index()
by_type.columns = ["Envelope Description", "Total Qty Purchased", "Total Receipt Amount", "Total Invoiced"]
total_purch = by_type["Total Qty Purchased"].sum()
by_type["% of Total Purchases"] = by_type["Total Qty Purchased"] / total_purch
by_type = by_type.sort_values("Total Qty Purchased", ascending=False).reset_index(drop=True)

# PART 4 - MONTHLY BY TYPE PIVOT
print("Building monthly-by-type pivot...")
pivot = ep.pivot_table(index="Month", columns="Envelope Description", values="Quantity Purchased", aggfunc="sum", fill_value=0)
col_order = pivot.sum().sort_values(ascending=False).index
pivot = pivot[col_order]
pivot = pivot.reset_index()
pivot["Month"] = pivot["Month"].dt.strftime("%b-%y")

# WRITE TO EXCEL
print("Writing Excel workbook...")
with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    monthly_out.to_excel(writer, sheet_name="Monthly Summary", index=False)
    annual.to_excel(writer, sheet_name="Annual Summary", index=False)
    by_type.to_excel(writer, sheet_name="By Envelope Type", index=False)
    pivot.to_excel(writer, sheet_name="Monthly by Type", index=False)

# FORMAT
print("Applying formatting...")
wb = load_workbook(OUT)
hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
data_font = Font(name="Calibri", size=11)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
red_font = Font(name="Calibri", size=11, color="CC0000")
green_font = Font(name="Calibri", size=11, color="006600")
COMMA_FMT = "#,##0"
CURRENCY_FMT = '$#,##0.00'
PCT_FMT = "0.0%"

def format_sheet(ws, col_formats, variance_cols=None):
    max_row = ws.max_row
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = data_font
            if c in col_formats and col_formats[c]:
                cell.number_format = col_formats[c]
    if variance_cols:
        for vc in variance_cols:
            col_letter = get_column_letter(vc)
            rng = f"{col_letter}2:{col_letter}{max_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], font=red_font))
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], font=green_font))
    for c in range(1, max_col + 1):
        header_len = len(str(ws.cell(row=1, column=c).value or ""))
        max_len = header_len
        for r in range(2, min(max_row + 1, 50)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                if isinstance(val, (int, float)):
                    max_len = max(max_len, len(f"{val:,.0f}"))
                else:
                    max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 35)
    ws.freeze_panes = "A2"

ws1 = wb["Monthly Summary"]
fmt1 = {2: COMMA_FMT, 3: COMMA_FMT, 4: COMMA_FMT, 5: COMMA_FMT, 6: COMMA_FMT, 7: COMMA_FMT, 8: COMMA_FMT, 9: COMMA_FMT}
format_sheet(ws1, fmt1, variance_cols=[6, 9])

ws2 = wb["Annual Summary"]
fmt2 = {2: COMMA_FMT, 3: COMMA_FMT, 4: COMMA_FMT, 5: COMMA_FMT, 6: COMMA_FMT, 7: PCT_FMT}
format_sheet(ws2, fmt2, variance_cols=[6])

ws3 = wb["By Envelope Type"]
fmt3 = {2: COMMA_FMT, 3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: PCT_FMT}
format_sheet(ws3, fmt3)

ws4 = wb["Monthly by Type"]
fmt4 = {c: COMMA_FMT for c in range(2, ws4.max_column + 1)}
format_sheet(ws4, fmt4)

wb.save(OUT)
print(f"\nWorkbook saved to:\n  {OUT}")

# CONSOLE SUMMARY
total_purchased = monthly["Purchased"].sum()
total_vol_usage = monthly["Volume Usage"].sum()
total_postal = monthly["Postal Usage"].sum()
total_spoilage = monthly["Spoilage"].sum()
net_variance = total_purchased - total_vol_usage
var_pct = net_variance / total_purchased * 100 if total_purchased else 0

print("\n" + "=" * 60)
print("  ENVELOPE RECONCILIATION SUMMARY  (Mar 2022 - Dec 2025)")
print("=" * 60)
print(f"  Total Envelopes Purchased:   {total_purchased:>14,}")
print(f"  Total Volume Usage:          {total_vol_usage:>14,}")
print(f"  Total Envelopes Mailed:      {total_postal:>14,}")
print(f"  Total Spoilage:              {total_spoilage:>14,}")
print(f"  Net Variance (Purch-Usage):  {net_variance:>14,}")
print(f"  Variance % of Purchases:     {var_pct:>13.1f}%")
print("=" * 60)
