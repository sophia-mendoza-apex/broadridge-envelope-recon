import os, re, sys, traceback
from datetime import datetime, date
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

BASE_PURCHASE = r'C:\Users\smendoza\OneDrive - Apex Clearing\Broadridge Billing\Billing Support\ICS Postage & Material Support\Print & Mail Reports\Envelope Purchase Orders'
BASE_BILLING  = r'C:\Users\smendoza\OneDrive - Apex Clearing\Broadridge Billing\Billing Support\ICS Postage & Material Support\Print & Mail Reports\Postage and Volume Report'
COMMODITIES_FILE = r'C:\Users\smendoza\OneDrive - Apex Clearing\Broadridge Billing\Billing Support\ICS Postage & Material Support\Print & Mail Reports\Commodities_Item_List Jan 2026.xlsx'
YTD_FILE = os.path.join(BASE_PURCHASE, 'Apex YTD Envelope Usage (2).xlsx')
OUTPUT_FILE = r'C:\Users\smendoza\Projects\Broadridge Envelopes\Envelope Reconciliation - Source Data.xlsx'
START_MONTH_KEY = "2020-01"  # Reconciliation starts January 2020
BASE_PRE2022_PURCHASE = os.path.join(os.path.dirname(OUTPUT_FILE), "Purchase Reports (from email)", "Purchase Reports")
BILLING_MASTER_2020 = os.path.join(os.path.dirname(OUTPUT_FILE), "Billing Master - 2020 - 2021.xlsx")
CONSOLIDATED_PURCHASES = os.path.join(BASE_PURCHASE, "2019-2022 Purchases.xlsx")


MONTH_NAMES = {"jan":1,"january":1,"feb":2,"february":2,"mar":3,"march":3,"apr":4,"april":4,"may":5,"jun":6,"june":6,"jul":7,"july":7,"aug":8,"august":8,"sep":9,"sept":9,"september":9,"oct":10,"october":10,"nov":11,"november":11,"dec":12,"december":12}
MONTH_ABBR = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

def safe_float(val, default=0.0):
    if val is None: return default
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip().replace(",","").replace("$","").replace(" ","")
    if s in ("","-","#VALUE!","#REF!","#N/A","N/A"): return default
    try: return float(s)
    except ValueError: return default

def safe_int(val, default=0): return int(safe_float(val, default))
def is_apex(cn):
    if cn is None: return False
    u = str(cn).upper()
    if "BROADRIDGE" in u: return False
    return "APEX" in u or "RIDGE" in u or "PENSON" in u or "PENSION" in u
def is_envelope(d): return d is not None and "ENV" in str(d).upper()

# Canonical envelope type mapping — consolidates variant descriptions from different
# source files (supplier IDs, verbose prefixes, revision dates) into 7 core types
# plus grouped categories for tax forms and non-Apex envelopes.
ENVELOPE_TYPE_RULES = [
    # Non-Apex envelopes (other Broadridge clients — exclude from totals)
    (r"ENVFID",   None),   # Fidelity
    (r"ENVHSBC",  None),   # HSBC
    (r"ENVMST",   None),   # Morgan Stanley

    # Core 7 envelope types (order matters — more specific patterns first)
    # Domestic Fold Statement (N14 PFC)
    (r"ENVAPXN14.*STMTPFC",        "ENVMEAPEXN14PFC"),
    (r"ENVMEAPEXN14PFC",            "ENVMEAPEXN14PFC"),

    # Domestic Flat Statement (9x12 PFC)
    (r"APEX.*9X12.*APEX9X12PFC",    "ENVMEAPEX9X12PFC"),
    (r"ENVMEAPEX9X12PFC",           "ENVMEAPEX9X12PFC"),

    # Foreign Fold Statement (N14 NI)
    (r"ENVAPXN14.*STMTNI",          "ENVMERIDGEN14NI11/08"),
    (r"ENVMERIDGEN14NI",            "ENVMERIDGEN14NI11/08"),

    # Foreign Flat Statement (9x12 NI)
    (r"ENVMERIDGE9X12NI",           "ENVMERIDGE9X12NI11/08"),

    # Domestic Confirms + Letters (#10 PFC) — all revisions
    (r"ENVAPXN10.*CNFPFC",          "ENVAPXN10 Confirms+Letters (PFC)"),
    (r"ENVAPXN10.*LTRPFC",          "ENVAPXN10 Confirms+Letters (PFC)"),
    (r"ENVAPXN10PFSCONN10",         "ENVAPXN10 Confirms+Letters (PFC)"),

    # Foreign Confirms + Letters (#10 NI)
    (r"ENVAPXN10.*CNFNI",           "ENVCONPFSN10NI"),
    (r"ENVCONPFSN10NI",             "ENVCONPFSN10NI"),
    (r"UNITED ENVELOPE.*ENVCONPFSN10NI", "ENVCONPFSN10NI"),

    # Flat Confirms (9x12 DW)
    (r"ENVCONRIDGE9X12DW",          "ENVCONRIDGE9X12DW"),

    # Tax form envelopes (group together)
    (r"ENV1099",                    "Tax Form Envelopes (1099/1099-R)"),
    (r"ENVAPEX.*1042",              "Tax Form Envelopes (1042/IRA)"),
]

def normalize_envelope_type(desc):
    """Map raw envelope description to canonical type. Returns None for non-Apex."""
    if not desc:
        return "(Unknown)"
    d = str(desc).strip().upper()
    for pattern, canonical in ENVELOPE_TYPE_RULES:
        if re.search(pattern, d, re.IGNORECASE):
            return canonical
    return str(desc).strip()  # Unknown — keep as-is

def parse_month_year_from_filename(fn):
    b = os.path.splitext(fn)[0]
    b = b.replace("Purchase Report ", "").replace("Purchase Report", "").replace("Copy of ", "").strip()
    m = re.search(r"(\d{1,2})-(\d{4})", b)
    if m: return int(m.group(1)), int(m.group(2))
    m = re.search(r"(\d{1,2})-(\d{2})", b)
    if m:
        mo, yr = int(m.group(1)), int(m.group(2))
        return mo, 2000+yr if yr<100 else yr
    return None, None

def parse_date_value(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    if isinstance(val, date): return datetime(val.year, val.month, val.day)
    if isinstance(val, (int, float)):
        try: return datetime(1899,12,30) + pd.Timedelta(days=int(val))
        except: return None
    s = str(val).strip()
    for fmt in ["%m/%d/%Y %H:%M:%S","%m/%d/%Y","%Y-%m-%d","%Y-%m-%d %H:%M:%S"]:
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

def make_month_key(m, y): return f"{y}-{m:02d}"

def month_key_to_label(k):
    p = k.split('-')
    mi = int(p[1])
    yi = int(p[0]) % 100
    return f"{MONTH_ABBR.get(mi, '???')}-{yi:02d}"

def open_workbook(fp):
    ext = os.path.splitext(fp)[1].lower()
    if ext == ".xls":
        # Legacy .xls format — convert to openpyxl-compatible via xlrd+openpyxl
        import xlrd
        xls = xlrd.open_workbook(fp)
        wb = openpyxl.Workbook()
        for si, sname in enumerate(xls.sheet_names()):
            xs = xls.sheet_by_name(sname)
            ws = wb.active if si == 0 else wb.create_sheet()
            ws.title = sname
            for row_idx in range(xs.nrows):
                for col_idx in range(xs.ncols):
                    cell = xs.cell(row_idx, col_idx)
                    val = cell.value
                    # Convert xlrd date numbers to datetime
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try: val = datetime(*xlrd.xldate_as_tuple(val, xls.datemode))
                        except: pass
                    ws.cell(row=row_idx+1, column=col_idx+1, value=val)
        return wb
    kw = {"data_only":True, "read_only":True}
    if ext == ".xlsm": kw["keep_vba"] = False
    return openpyxl.load_workbook(fp, **kw)

purchase_records = []
volume_records = []
postage_records = []
ytd_usage_records = []
commodities_records = []
data_quality_issues = []
files_processed_purchase = 0
files_failed_purchase = 0
files_processed_billing = 0
files_failed_billing = 0


def read_consolidated_purchases():
    """Read 2019-2022 Purchases.xlsx — consolidated PO history with PO numbers."""
    global files_processed_purchase, files_failed_purchase
    if not os.path.exists(CONSOLIDATED_PURCHASES):
        return
    try:
        wb = open_workbook(CONSOLIDATED_PURCHASES)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        headers = [str(c).strip() if c else '' for c in rows[0]]
        hmap = {h: i for i, h in enumerate(headers) if h}
        count = 0
        for r in rows[1:]:
            vals = list(r)
            if len(vals) < 5: continue
            desc = str(vals[hmap.get("WMS Item #", 0)] or "").strip()
            if not is_envelope(desc): continue
            po_date = parse_date_value(vals[hmap.get("PO Date", 3)])
            if po_date is None: continue
            qty_raw = safe_float(vals[hmap.get("Quantity Ordered", 4)])
            uom = str(vals[hmap.get("UOM", 5)] or "").strip().upper()
            qty = qty_raw * 1000 if uom in ("M", "TH") else qty_raw
            unit_price = safe_float(vals[hmap.get("Unit Price", 6)])
            qty_recv = safe_float(vals[hmap.get("Quantity Received", 7)]) * (1000 if uom in ("M", "TH") else 1)
            total_cost = safe_float(vals[hmap.get("Total Amount", 8)])
            po_number = str(vals[hmap.get("PO Number", 2)] or "").strip()
            mk = make_month_key(po_date.month, po_date.year)
            purchase_records.append({
                "month_key": mk, "po_date": po_date, "client": "APEX",
                "description": desc, "qty_ordered": qty, "qty_received": qty_recv,
                "unit_price": unit_price, "total_cost": total_cost,
                "uom": uom, "source_file": "2019-2022 Purchases.xlsx",
                "po_number": po_number, "markup_pct": 0.0, "invoiced_amount": total_cost
            })
            count += 1
        files_processed_purchase += 1
        print(f"    Consolidated 2019-2022: {count} records")
    except Exception as e:
        files_failed_purchase += 1
        data_quality_issues.append(("Purchase", "2019-2022 Purchases.xlsx", f"Error: {e}"))
        traceback.print_exc()

def read_apex_2022_purchases():
    global files_processed_purchase, files_failed_purchase
    fp = os.path.join(BASE_PURCHASE, 'Apex 2022 Purchases.xlsx')
    if not os.path.exists(fp):
        data_quality_issues.append(("Purchase", "2022", "Apex 2022 Purchases.xlsx not found"))
        return
    try:
        wb = open_workbook(fp)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        headers = [str(c).strip() if c else '' for c in rows[0]]
        hmap = {h:i for i,h in enumerate(headers) if h}
        for r in rows[1:]:
            vals = list(r)
            if len(vals) < 5: continue
            client = str(vals[hmap.get("Client Name",1)] or "").strip()
            if not is_apex(client): continue
            desc = str(vals[hmap.get("Item Description",4)] or "").strip()
            if not is_envelope(desc): continue
            po_date = parse_date_value(vals[hmap.get("PO Date",0)])
            if po_date is None: continue
            qty_raw = safe_float(vals[hmap.get("Quantity",6)])
            uom = str(vals[hmap.get("UOM",7)] or "").strip().upper()
            qty = qty_raw * 1000 if uom in ("M", "TH") else qty_raw
            unit_price = safe_float(vals[hmap.get("Unit Price",8)])
            total_cost = safe_float(vals[hmap.get("Markup%",9)])
            if total_cost == 0: total_cost = unit_price * qty_raw
            mk = make_month_key(po_date.month, po_date.year)
            purchase_records.append({
                "month_key": mk, "po_date": po_date, "client": client,
                "description": desc, "qty_ordered": qty, "qty_received": qty,
                "unit_price": unit_price, "total_cost": total_cost,
                "uom": uom, "source_file": "Apex 2022 Purchases.xlsx",
                "po_number": "", "markup_pct": 0.0, "invoiced_amount": total_cost
            })
        files_processed_purchase += 1
    except Exception as e:
        files_failed_purchase += 1
        data_quality_issues.append(("Purchase", "2022", f"Error reading Apex 2022 Purchases: {e}"))
        traceback.print_exc()

def read_standard_purchase(fp, fn):
    global files_processed_purchase, files_failed_purchase
    try:
        wb = open_workbook(fp)
        sn = "Final Data" if "Final Data" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows: return
        headers = [str(c).strip() if c else '' for c in rows[0]]
        hmap = {h:i for i,h in enumerate(headers) if h}
        mo, yr = parse_month_year_from_filename(fn)
        count = 0
        for r in rows[1:]:
            vals = list(r)
            if len(vals) < 5: continue
            ci = hmap.get("Client Name", hmap.get("Client", -1))
            if ci >= 0 and ci < len(vals):
                client = str(vals[ci] or '').strip()
                if not is_apex(client): continue
            else:
                # No client column — check description/item fields for APEX
                row_text = " ".join(str(v or '') for v in vals).upper()
                if "APEX" not in row_text and "RIDGE" not in row_text: continue
                if "BROADRIDGE" in row_text and "APEX" not in row_text: continue
                client = "APEX"
            di = hmap.get("Item Description", hmap.get("Description", -1))
            if di < 0 or di >= len(vals): continue
            desc = str(vals[di] or '').strip()
            if not is_envelope(desc): continue
            pdi = hmap.get("PO Date", hmap.get("Receipt Date", hmap.get("Order Date", -1)))
            po_date = parse_date_value(vals[pdi]) if pdi >= 0 and pdi < len(vals) else None
            if po_date is None and mo and yr: po_date = datetime(yr, mo, 15)
            if po_date is None: continue
            qi = hmap.get("Quantity", hmap.get("Qty", -1))
            qty_raw = safe_float(vals[qi]) if qi >= 0 and qi < len(vals) else 0
            ui = hmap.get("UOM", -1)
            uom = str(vals[ui] or '').strip().upper() if ui >= 0 and ui < len(vals) else ''
            qty = qty_raw * 1000 if uom in ("M", "TH") else qty_raw
            qri = hmap.get("Quantity Received", hmap.get("Qty Recd.", -1))
            qty_rcv = safe_float(vals[qri]) if qri >= 0 and qri < len(vals) else qty_raw
            qty_rcv_units = qty_rcv * 1000 if uom in ("M", "TH") else qty_rcv
            upi = hmap.get("Unit Price", hmap.get("Rate", -1))
            unit_price = safe_float(vals[upi]) if upi >= 0 and upi < len(vals) else 0
            rai = hmap.get("Receipt Amount", hmap.get("Total Price", -1))
            receipt_amt = safe_float(vals[rai]) if rai >= 0 and rai < len(vals) else 0
            mai = hmap.get("Mark up %", hmap.get("Markup", hmap.get("Mark % 1", -1)))
            markup_val = safe_float(vals[mai]) if mai >= 0 and mai < len(vals) else 0
            invoiced = markup_val if markup_val > 100 else receipt_amt
            total_cost = invoiced if invoiced > 0 else receipt_amt
            if total_cost == 0: total_cost = unit_price * qty_raw
            poi = hmap.get("PO Number", hmap.get("Order ID", -1))
            po_num = str(vals[poi] or '').strip() if poi >= 0 and poi < len(vals) else ''
            mk = make_month_key(po_date.month, po_date.year)
            purchase_records.append({
                "month_key": mk, "po_date": po_date, "client": client,
                "description": desc, "qty_ordered": qty, "qty_received": qty_rcv_units,
                "unit_price": unit_price, "total_cost": total_cost,
                "uom": uom, "source_file": fn,
                "po_number": po_num, "markup_pct": 0.0, "invoiced_amount": invoiced
            })
            count += 1
        files_processed_purchase += 1
        if count == 0: data_quality_issues.append(("Purchase", fn, "No APEX envelope rows found"))
    except Exception as e:
        files_failed_purchase += 1
        data_quality_issues.append(("Purchase", fn, f"Error: {e}"))
        traceback.print_exc()

def read_new_format_purchase(fp, fn):
    global files_processed_purchase, files_failed_purchase
    try:
        wb = open_workbook(fp)
        sn = "Purchase Report" if "Purchase Report" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows: return
        headers = [str(c).strip() if c else '' for c in rows[0]]
        hmap = {h:i for i,h in enumerate(headers) if h}
        mo, yr = parse_month_year_from_filename(fn)
        count = 0
        for r in rows[1:]:
            vals = list(r)
            if len(vals) < 10: continue
            ci = hmap.get("Client Name", hmap.get("Client", -1))
            if ci < 0 or ci >= len(vals): continue
            client = str(vals[ci] or '').strip()
            if not is_apex(client): continue
            di = hmap.get("Description", -1)
            desc_val = str(vals[di] or '').strip() if di >= 0 and di < len(vals) else ''
            wms_i = hmap.get("WMS Item Number", -1)
            wms_val = str(vals[wms_i] or '').strip() if wms_i >= 0 and wms_i < len(vals) else ''
            if not is_envelope(desc_val) and not is_envelope(wms_val): continue
            desc = wms_val if is_envelope(wms_val) else desc_val
            rdi = hmap.get("Receipt Date", hmap.get("Creation Date", -1))
            po_date = parse_date_value(vals[rdi]) if rdi >= 0 and rdi < len(vals) else None
            if po_date is None and mo and yr: po_date = datetime(yr, mo, 15)
            if po_date is None: continue
            oqi = hmap.get("Ordered Qty", -1)
            ord_raw = safe_float(vals[oqi]) if oqi >= 0 and oqi < len(vals) else 0
            rqi = hmap.get("Received Qty", -1)
            rcv_raw = safe_float(vals[rqi]) if rqi >= 0 and rqi < len(vals) else ord_raw
            # In new format, Ordered Qty is already in individual units
            # UOM is the pricing basis (e.g., 1000 = price per 1000 units)
            # So we do NOT multiply qty by UOM
            ui = hmap.get("UOM", -1)
            uom_val = safe_float(vals[ui]) if ui >= 0 and ui < len(vals) else 1
            qty_ordered = ord_raw
            qty_received = rcv_raw
            uom_str = "M" if uom_val >= 100 else "EA"
            uci = hmap.get("Unit Cost", -1)
            unit_cost = safe_float(vals[uci]) if uci >= 0 and uci < len(vals) else 0
            tci = hmap.get("Total Cost", -1)
            total_cost = safe_float(vals[tci]) if tci >= 0 and tci < len(vals) else 0
            if total_cost == 0: total_cost = unit_cost * (ord_raw / uom_val) if uom_val > 0 else 0
            mpi = hmap.get("Markup %", -1)
            markup_pct = safe_float(vals[mpi]) if mpi >= 0 and mpi < len(vals) else 0
            mti = hmap.get("Markup Total", -1)
            markup_total = safe_float(vals[mti]) if mti >= 0 and mti < len(vals) else 0
            invoiced = total_cost + markup_total if markup_total > 0 else total_cost * (1 + markup_pct)
            poi = hmap.get("PO Number", -1)
            po_num = str(vals[poi] or '').strip() if poi >= 0 and poi < len(vals) else ''
            mk = make_month_key(po_date.month, po_date.year)
            purchase_records.append({
                "month_key": mk, "po_date": po_date, "client": client,
                "description": desc, "qty_ordered": qty_ordered, "qty_received": qty_received,
                "unit_price": unit_cost, "total_cost": total_cost,
                "uom": uom_str, "source_file": fn,
                "po_number": po_num, "markup_pct": markup_pct, "invoiced_amount": invoiced
            })
            count += 1
        files_processed_purchase += 1
        if count == 0: data_quality_issues.append(("Purchase", fn, "No APEX envelope rows found"))
    except Exception as e:
        files_failed_purchase += 1
        data_quality_issues.append(("Purchase", fn, f"Error: {e}"))
        traceback.print_exc()

def process_pre2022_purchase_reports():
    """Walk the nested FY'20 and FY'21 directories for purchase reports."""
    if not os.path.isdir(BASE_PRE2022_PURCHASE):
        data_quality_issues.append(("Purchase", "Pre-2022", f"Directory not found: {BASE_PRE2022_PURCHASE}"))
        return
    for fy_folder in ["FY'20", "FY'21"]:
        fy_path = os.path.join(BASE_PRE2022_PURCHASE, fy_folder)
        if not os.path.isdir(fy_path): continue
        for sub in sorted(os.listdir(fy_path)):
            sub_path = os.path.join(fy_path, sub)
            if not os.path.isdir(sub_path): continue
            for fn in os.listdir(sub_path):
                fl = fn.lower()
                if not (fl.endswith(".xlsx") or fl.endswith(".xlsm") or fl.endswith(".xls")): continue
                fp = os.path.join(sub_path, fn)
                read_standard_purchase(fp, fn)

def process_all_purchase_reports():
    read_consolidated_purchases()
    process_pre2022_purchase_reports()
    read_apex_2022_purchases()
    for fn in os.listdir(BASE_PURCHASE):
        fl = fn.lower()
        if not fl.startswith("purchase report"): continue
        if not (fl.endswith(".xlsx") or fl.endswith(".xlsm") or fl.endswith(".xls")): continue
        if fn.startswith("Copy of "): continue
        fp = os.path.join(BASE_PURCHASE, fn)
        mo, yr = parse_month_year_from_filename(fn)
        if fl.endswith(".xlsx") and yr and yr >= 2025:
            read_new_format_purchase(fp, fn)
        else:
            read_standard_purchase(fp, fn)


def read_billing_workbook(fp, fn, year_folder):
    global files_processed_billing, files_failed_billing
    try:
        wb = open_workbook(fp)
        sheet_names = wb.sheetnames
        # Build lookup: strip whitespace from sheet names for flexible matching
        sheet_lookup = {sn.strip().lower(): sn for sn in sheet_names}
        # Read Volume Data (handle "Volume Data", "Volume ", "Volume", etc.)
        vol_sheet = None
        for candidate in ["volume data", "volume"]:
            if candidate in sheet_lookup:
                vol_sheet = sheet_lookup[candidate]
                break
        if vol_sheet:
            ws = wb[vol_sheet]
            all_rows = list(ws.iter_rows(values_only=True))
            header_map = None
            data_start = 0
            for i, r in enumerate(all_rows):
                vals = list(r)
                if vals and vals[0] is not None and str(vals[0]).strip() == 'Job_Index':
                    hm = {}
                    for j, v in enumerate(vals):
                        if v is not None:
                            key = str(v).strip()
                            if key not in hm:
                                hm[key] = j
                    if 'Envelopes' in hm:
                        header_map = hm
                        data_start = i + 1
                elif header_map is not None and vals and vals[0] is not None:
                    if str(vals[0]).strip() != 'Job_Index':
                        break
            if header_map is None:
                data_quality_issues.append(("Billing-Vol", fn, "No Volume Data headers found"))
            else:
                for r in all_rows[data_start:]:
                    vals = list(r)
                    if not vals or vals[0] is None: continue
                    if str(vals[0]).strip() == 'Job_Index': continue
                    cn_i = header_map.get("Client_Name", header_map.get("Client Name", -1))
                    if cn_i < 0 or cn_i >= len(vals): continue
                    client = str(vals[cn_i] or '').strip()
                    if not is_apex(client): continue
                    bm_i = header_map.get("Billing_Month", -1)
                    by_i = header_map.get("Billing_Year", -1)
                    bm = safe_int(vals[bm_i]) if bm_i >= 0 and bm_i < len(vals) else 0
                    by = safe_int(vals[by_i]) if by_i >= 0 and by_i < len(vals) else 0
                    if bm == 0 or by == 0: continue
                    env_i = header_map.get("Envelopes", -1)
                    envelopes = safe_int(vals[env_i]) if env_i >= 0 and env_i < len(vals) else 0
                    pn_i = header_map.get("Product_Name", header_map.get("Product Name", -1))
                    prod_name = str(vals[pn_i] or '').strip() if pn_i >= 0 and pn_i < len(vals) else ''
                    pc_i = header_map.get("Product_Category", header_map.get("Product Category", -1))
                    prod_cat = str(vals[pc_i] or '').strip() if pc_i >= 0 and pc_i < len(vals) else ''
                    mk = make_month_key(bm, by)
                    volume_records.append({
                        "month_key": mk, "billing_month": bm, "billing_year": by,
                        "client": client, "envelopes": envelopes, "images": 0,
                        "sheets": 0, "product_name": prod_name,
                        "product_category": prod_cat, "source_file": fn
                    })

        # Read Postage Data (handle "Postage Data", "Postage", etc.)
        post_sheet = None
        for candidate in ["postage data", "postage"]:
            if candidate in sheet_lookup:
                post_sheet = sheet_lookup[candidate]
                break
        if post_sheet:
            ws = wb[post_sheet]
            all_rows = list(ws.iter_rows(values_only=True))
            header_map = None
            data_start = 0
            for i, r in enumerate(all_rows):
                vals = list(r)
                if vals and vals[0] is not None and str(vals[0]).strip() == 'Job_Index':
                    hm = {}
                    for j, v in enumerate(vals):
                        if v is not None:
                            hm[str(v).strip()] = j
                    if 'Env_Mailed' in hm and 'Billing_Month' in hm:
                        header_map = hm
                        data_start = i + 1
                elif header_map is not None and vals and vals[0] is not None:
                    if str(vals[0]).strip() != 'Job_Index':
                        break
            if header_map is None:
                data_quality_issues.append(("Billing-Post", fn, "No Postage headers with Env_Mailed found"))
            else:
                for r in all_rows[data_start:]:
                    vals = list(r)
                    if not vals or vals[0] is None: continue
                    if str(vals[0]).strip() == 'Job_Index': continue
                    cn_i = header_map.get("Client_Name", header_map.get("Client Name", -1))
                    if cn_i < 0 or cn_i >= len(vals): continue
                    client = str(vals[cn_i] or '').strip()
                    if not is_apex(client): continue
                    bm_i = header_map.get("Billing_Month", -1)
                    by_i = header_map.get("Billing_Year", -1)
                    bm = safe_int(vals[bm_i]) if bm_i >= 0 and bm_i < len(vals) else 0
                    by = safe_int(vals[by_i]) if by_i >= 0 and by_i < len(vals) else 0
                    if bm == 0 or by == 0: continue
                    em_i = header_map.get("Env_Mailed", -1)
                    env_mailed = safe_int(vals[em_i]) if em_i >= 0 and em_i < len(vals) else 0
                    sp_i = header_map.get("Spoils", -1)
                    spoils = safe_int(vals[sp_i]) if sp_i >= 0 and sp_i < len(vals) else 0
                    mk = make_month_key(bm, by)
                    postage_records.append({
                        "month_key": mk, "billing_month": bm, "billing_year": by,
                        "client": client, "env_mailed": env_mailed, "spoils": spoils,
                        "source_file": fn
                    })

        wb.close()
        files_processed_billing += 1
    except Exception as e:
        files_failed_billing += 1
        data_quality_issues.append(("Billing", fn, f"Error: {e}"))
        traceback.print_exc()

def is_billing_duplicate(fn):
    fl = fn.lower()
    if "paper usage" in fl: return True
    if " (1)" in fn: return True
    if "billing master" in fl: return True
    return False

def parse_billing_month_from_filename(fn, year_folder):
    fl = fn.lower()
    m = re.match(r'^(\d{1,2})\s', fn)
    if m:
        return int(m.group(1)), int(year_folder)
    for mname, mnum in MONTH_NAMES.items():
        if mname in fl:
            return mnum, int(year_folder)
    return None, None

def process_all_billing_workbooks():
    # Process Billing Master 2020-2021 (single file with Volume Data + Postage Data for all of 2020-2021)
    if os.path.exists(BILLING_MASTER_2020):
        read_billing_workbook(BILLING_MASTER_2020, os.path.basename(BILLING_MASTER_2020), "2020-2021")

    for yr_folder in ['2022','2023','2024','2025']:
        yr_path = os.path.join(BASE_BILLING, yr_folder)
        if not os.path.isdir(yr_path): continue
        files = sorted(os.listdir(yr_path))

        # Separate master vs non-master files; skip paper usage and (1) copies
        master_files = []
        non_master_files = []
        for fn in files:
            fl = fn.lower()
            if not fl.endswith(".xlsx"): continue
            if "paper usage" in fl: continue
            if " (1)" in fn: continue
            if "billing master" in fl:
                master_files.append(fn)
            else:
                non_master_files.append(fn)

        # Find which months are covered by non-master files
        covered_months = set()
        for fn in non_master_files:
            mo, _ = parse_billing_month_from_filename(fn, yr_folder)
            if mo: covered_months.add(mo)

        # Process non-master files first
        for fn in non_master_files:
            fp = os.path.join(yr_path, fn)
            read_billing_workbook(fp, fn, yr_folder)

        # Process master files only when no non-master exists for that month
        for fn in master_files:
            mo, _ = parse_billing_month_from_filename(fn, yr_folder)
            if mo and mo in covered_months:
                continue
            fp = os.path.join(yr_path, fn)
            read_billing_workbook(fp, fn, yr_folder)


def read_ytd_usage():
    if not os.path.exists(YTD_FILE):
        data_quality_issues.append(("YTD", "N/A", "YTD file not found"))
        return
    try:
        wb = open_workbook(YTD_FILE)
        month_tabs = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug']
        abbr_values = list(MONTH_ABBR.values())
        for tab_name in month_tabs:
            if tab_name not in wb.sheetnames: continue
            ws = wb[tab_name]
            rows = list(ws.iter_rows(values_only=True))
            mi = abbr_values.index(tab_name) + 1 if tab_name in abbr_values else 0
            if mi == 0: continue
            for r in rows:
                vals = list(r)
                if len(vals) < 2: continue
                item = str(vals[0] or '').strip()
                if not item or item.lower() == 'materials': continue
                if not is_envelope(item): continue
                usage = safe_int(vals[1])
                if usage == 0: continue
                mk = make_month_key(mi, 2022)
                ytd_usage_records.append({
                    "month_key": mk, "item": item, "usage": usage,
                    "source": "YTD-Monthly", "year": 2022
                })

        if "Apex Billing Sheet" in wb.sheetnames:
            ws = wb["Apex Billing Sheet"]
            rows = list(ws.iter_rows(values_only=True))
            for r in rows:
                vals = list(r)
                if len(vals) < 2: continue
                item = str(vals[0] or '').strip()
                if not item: continue
                if not is_envelope(item): continue
                usage = safe_int(vals[1])
                if usage == 0: continue
                ytd_usage_records.append({
                    "month_key": "YTD-2022", "item": item, "usage": usage,
                    "source": "YTD-Summary", "year": 2022
                })

        if "Volume Data" in wb.sheetnames:
            ws = wb["Volume Data"]
            all_rows = list(ws.iter_rows(values_only=True))
            header_map = None
            data_start = 0
            for i, r in enumerate(all_rows):
                vals = list(r)
                if vals and vals[0] is not None and str(vals[0]).strip() == 'Job_Index':
                    hm = {}
                    for j, v in enumerate(vals):
                        if v is not None:
                            hm[str(v).strip()] = j
                    if 'Envelopes' in hm:
                        header_map = hm
                        data_start = i + 1
                elif header_map is not None and vals and vals[0] is not None:
                    if str(vals[0]).strip() != 'Job_Index':
                        break
            if header_map:
                for r in all_rows[data_start:]:
                    vals = list(r)
                    if not vals or vals[0] is None: continue
                    if str(vals[0]).strip() == 'Job_Index': continue
                    cn_i = header_map.get("Client_Name", header_map.get("Client Name", -1))
                    if cn_i < 0 or cn_i >= len(vals): continue
                    client = str(vals[cn_i] or '').strip()
                    if not is_apex(client): continue
                    bm_i = header_map.get("Billing_Month", -1)
                    by_i = header_map.get("Billing_Year", -1)
                    bm = safe_int(vals[bm_i]) if bm_i >= 0 and bm_i < len(vals) else 0
                    by = safe_int(vals[by_i]) if by_i >= 0 and by_i < len(vals) else 0
                    if bm == 0 or by == 0: continue
                    env_i = header_map.get("Envelopes", -1)
                    envelopes = safe_int(vals[env_i]) if env_i >= 0 and env_i < len(vals) else 0
                    mk = make_month_key(bm, by)
                    volume_records.append({
                        "month_key": mk, "billing_month": bm, "billing_year": by,
                        "client": client, "envelopes": envelopes, "images": 0,
                        "sheets": 0, "product_name": "",
                        "product_category": "", "source_file": "YTD File - Volume Data"
                    })

        if "Postage Data" in wb.sheetnames:
            ws = wb["Postage Data"]
            all_rows = list(ws.iter_rows(values_only=True))
            header_map = None
            data_start = 0
            for i, r in enumerate(all_rows):
                vals = list(r)
                if vals and vals[0] is not None and str(vals[0]).strip() == 'Job_Index':
                    hm = {}
                    for j, v in enumerate(vals):
                        if v is not None:
                            hm[str(v).strip()] = j
                    if 'Env_Mailed' in hm and 'Billing_Month' in hm:
                        header_map = hm
                        data_start = i + 1
                elif header_map is not None and vals and vals[0] is not None:
                    if str(vals[0]).strip() != 'Job_Index':
                        break
            if header_map:
                for r in all_rows[data_start:]:
                    vals = list(r)
                    if not vals or vals[0] is None: continue
                    if str(vals[0]).strip() == 'Job_Index': continue
                    cn_i = header_map.get("Client_Name", header_map.get("Client Name", -1))
                    if cn_i < 0 or cn_i >= len(vals): continue
                    client = str(vals[cn_i] or '').strip()
                    if not is_apex(client): continue
                    bm_i = header_map.get("Billing_Month", -1)
                    by_i = header_map.get("Billing_Year", -1)
                    bm = safe_int(vals[bm_i]) if bm_i >= 0 and bm_i < len(vals) else 0
                    by = safe_int(vals[by_i]) if by_i >= 0 and by_i < len(vals) else 0
                    if bm == 0 or by == 0: continue
                    em_i = header_map.get("Env_Mailed", -1)
                    env_mailed = safe_int(vals[em_i]) if em_i >= 0 and em_i < len(vals) else 0
                    sp_i = header_map.get("Spoils", -1)
                    spoils = safe_int(vals[sp_i]) if sp_i >= 0 and sp_i < len(vals) else 0
                    mk = make_month_key(bm, by)
                    postage_records.append({
                        "month_key": mk, "billing_month": bm, "billing_year": by,
                        "client": client, "env_mailed": env_mailed, "spoils": spoils,
                        "source_file": "YTD File - Postage Data"
                    })

        wb.close()
    except Exception as e:
        data_quality_issues.append(("YTD", "N/A", f"Error: {e}"))
        traceback.print_exc()

def read_commodities():
    if not os.path.exists(COMMODITIES_FILE):
        data_quality_issues.append(("Commodities", "N/A", "Commodities file not found"))
        return
    try:
        wb = open_workbook(COMMODITIES_FILE)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 3: return
        headers = [str(c).strip() if c else '' for c in rows[1]]
        hmap = {h:i for i,h in enumerate(headers) if h}
        for r in rows[2:]:
            vals = list(r)
            if len(vals) < 5: continue
            item_i = hmap.get("Item Number", 0)
            item = str(vals[item_i] or '').strip()
            if not item: continue
            if not is_envelope(item): continue
            client_i = hmap.get("Client", 4)
            client = str(vals[client_i] or '').strip()
            if not is_apex(client): continue
            desc_i = hmap.get("Description", 2)
            desc = str(vals[desc_i] or '').strip()
            inv_i = hmap.get("Total Inventory", 13)
            total_inv = safe_int(vals[inv_i]) if inv_i < len(vals) else 0
            locs = {}
            loc_names = ["Edgewood Inventory","South Windsor Inventory",
                         "Coppell Inventory","Kansas City Inventory",
                         "El Dorado Hills Inventory","Canada Inventory"]
            for loc_name in loc_names:
                li = hmap.get(loc_name, -1)
                if li >= 0 and li < len(vals):
                    short = loc_name.replace(' Inventory','')
                    locs[short] = safe_int(vals[li])
            size_i = hmap.get("Size / Month", 5)
            size = str(vals[size_i] or '').strip() if size_i < len(vals) else ''
            commodities_records.append({
                "item_number": item, "description": desc, "client": client,
                "size": size, "total_inventory": total_inv, "locations": locs
            })
    except Exception as e:
        data_quality_issues.append(("Commodities", "N/A", f"Error: {e}"))
        traceback.print_exc()


def build_output():
    wb = openpyxl.Workbook()

    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_fmt_int = "#,##0"
    num_fmt_currency = "#,##0.00"
    num_fmt_pct = "0.0%"
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = ws.cell(row=row+1, column=1).coordinate

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Aggregate data
    monthly_purchased = defaultdict(float)
    monthly_cost = defaultdict(float)
    monthly_invoiced = defaultdict(float)
    for rec in purchase_records:
        if normalize_envelope_type(rec["description"]) is None:
            continue  # Exclude non-Apex envelopes from totals
        mk = rec["month_key"]
        monthly_purchased[mk] += rec["qty_ordered"]
        monthly_cost[mk] += rec["total_cost"]
        monthly_invoiced[mk] += rec["invoiced_amount"]

    monthly_envelopes_used = defaultdict(float)
    for rec in volume_records:
        mk = rec["month_key"]
        monthly_envelopes_used[mk] += rec["envelopes"]

    monthly_mailed = defaultdict(float)
    monthly_spoils = defaultdict(float)
    for rec in postage_records:
        mk = rec["month_key"]
        monthly_mailed[mk] += rec["env_mailed"]
        monthly_spoils[mk] += rec["spoils"]

    all_keys = sorted(k for k in set(
        list(monthly_purchased.keys()) +
        list(monthly_envelopes_used.keys()) +
        list(monthly_mailed.keys())
    ) if k >= START_MONTH_KEY)

    # === TAB 1: Monthly Summary ===
    ws1 = wb.active
    ws1.title = "Monthly Summary"
    headers1 = ["Month", "Envelopes Purchased", "Envelopes Used (Volume)",
                "Envelopes Mailed (Postage)", "Spoils",
                "Net Variance (Purchased - Used)", "Variance %",
                "Purchase Cost", "Invoiced Amount"]
    style_header(ws1, headers1)
    for i, mk in enumerate(all_keys, 2):
        label = month_key_to_label(mk)
        purchased = monthly_purchased.get(mk, 0)
        used = monthly_envelopes_used.get(mk, 0)
        mailed = monthly_mailed.get(mk, 0)
        spoils = monthly_spoils.get(mk, 0)
        variance = purchased - used
        var_pct = variance / purchased if purchased > 0 else 0
        cost = monthly_cost.get(mk, 0)
        invoiced = monthly_invoiced.get(mk, 0)
        row_data = [label, purchased, used, mailed, spoils, variance, var_pct, cost, invoiced]
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=c, value=val)
            cell.border = thin_border
            if c in (2,3,4,5,6): cell.number_format = num_fmt_int
            elif c == 7: cell.number_format = num_fmt_pct
            elif c in (8,9): cell.number_format = num_fmt_currency
    if len(all_keys) > 0:
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100")
        rng = f"F2:F{len(all_keys)+1}"
        ws1.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font))
        ws1.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill, font=green_font))
    auto_width(ws1)

    # === TAB 2: Annual Summary ===
    ws2 = wb.create_sheet("Annual Summary")
    headers2 = ["Year", "Envelopes Purchased", "Envelopes Used (Volume)",
                "Envelopes Mailed (Postage)", "Spoils",
                "Net Variance", "Variance %", "Total Cost", "Total Invoiced"]
    style_header(ws2, headers2)
    yearly = defaultdict(lambda: [0,0,0,0,0,0])
    for mk in all_keys:
        yr = int(mk.split('-')[0])
        yearly[yr][0] += monthly_purchased.get(mk, 0)
        yearly[yr][1] += monthly_envelopes_used.get(mk, 0)
        yearly[yr][2] += monthly_mailed.get(mk, 0)
        yearly[yr][3] += monthly_spoils.get(mk, 0)
        yearly[yr][4] += monthly_cost.get(mk, 0)
        yearly[yr][5] += monthly_invoiced.get(mk, 0)
    for i, yr in enumerate(sorted(yearly.keys()), 2):
        d = yearly[yr]
        variance = d[0] - d[1]
        var_pct = variance / d[0] if d[0] > 0 else 0
        row_data = [yr, d[0], d[1], d[2], d[3], variance, var_pct, d[4], d[5]]
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=c, value=val)
            cell.border = thin_border
            if c in (2,3,4,5,6): cell.number_format = num_fmt_int
            elif c == 7: cell.number_format = num_fmt_pct
            elif c in (8,9): cell.number_format = num_fmt_currency
    auto_width(ws2)

    # === TAB 3: By Envelope Type ===
    ws3 = wb.create_sheet("By Envelope Type")
    headers3 = ["Envelope Type", "Total Purchased", "Total Cost", "Avg Unit Price",
                "First Purchase", "Last Purchase"]
    style_header(ws3, headers3)
    by_type = defaultdict(lambda: [0, 0.0, [], None, None])
    excluded_non_apex = 0
    for rec in purchase_records:
        if rec["month_key"] < START_MONTH_KEY: continue
        d = normalize_envelope_type(rec["description"])
        if d is None:
            excluded_non_apex += rec["qty_ordered"]
            continue  # Non-Apex envelope (Fidelity, HSBC, Morgan Stanley)
        by_type[d][0] += rec["qty_ordered"]
        by_type[d][1] += rec["total_cost"]
        by_type[d][2].append(rec["unit_price"])
        pd_val = rec["po_date"]
        if by_type[d][3] is None or pd_val < by_type[d][3]: by_type[d][3] = pd_val
        if by_type[d][4] is None or pd_val > by_type[d][4]: by_type[d][4] = pd_val
    if excluded_non_apex:
        print(f"  Non-Apex envelopes excluded from By Type: {excluded_non_apex:,}")
    for i, (etype, vals) in enumerate(sorted(by_type.items()), 2):
        avg_price = sum(vals[2]) / len(vals[2]) if vals[2] else 0
        first_dt = vals[3].strftime("%m/%d/%Y") if vals[3] else ""
        last_dt = vals[4].strftime("%m/%d/%Y") if vals[4] else ""
        row_data = [etype, vals[0], vals[1], avg_price, first_dt, last_dt]
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=i, column=c, value=val)
            cell.border = thin_border
            if c == 2: cell.number_format = num_fmt_int
            elif c in (3,4): cell.number_format = num_fmt_currency
    auto_width(ws3)

    # === TAB 3b: Usage by Product ===
    ws3b = wb.create_sheet("Usage by Product")
    headers3b = ["Product Name", "Total Envelopes Used", "First Month", "Last Month"]
    style_header(ws3b, headers3b)
    by_product = defaultdict(lambda: [0, None, None])
    for rec in volume_records:
        if rec["month_key"] < START_MONTH_KEY: continue
        pn = rec.get("product_name", "").strip() or "(Unknown)"
        by_product[pn][0] += rec["envelopes"]
        mk = rec["month_key"]
        if by_product[pn][1] is None or mk < by_product[pn][1]: by_product[pn][1] = mk
        if by_product[pn][2] is None or mk > by_product[pn][2]: by_product[pn][2] = mk
    for i, (pname, vals) in enumerate(sorted(by_product.items(), key=lambda x: -x[1][0]), 2):
        first_m = month_key_to_label(vals[1]) if vals[1] else ""
        last_m = month_key_to_label(vals[2]) if vals[2] else ""
        row_data = [pname, vals[0], first_m, last_m]
        for c, val in enumerate(row_data, 1):
            cell = ws3b.cell(row=i, column=c, value=val)
            cell.border = thin_border
            if c == 2: cell.number_format = num_fmt_int
    auto_width(ws3b)

    # === TAB 4: Purchase Detail ===
    ws4 = wb.create_sheet("Purchase Detail")
    headers4 = ["Month", "PO Number", "PO Date", "Client", "Description",
                "Qty Ordered", "Qty Received", "UOM", "Unit Price",
                "Total Cost", "Invoiced Amount", "Source File"]
    style_header(ws4, headers4)
    sorted_purchases = sorted((r for r in purchase_records if r["month_key"] >= START_MONTH_KEY and normalize_envelope_type(r["description"]) is not None), key=lambda x: x["month_key"])
    for i, rec in enumerate(sorted_purchases, 2):
        po_dt = rec["po_date"].strftime("%m/%d/%Y") if rec["po_date"] else ""
        row_data = [
            month_key_to_label(rec["month_key"]),
            rec["po_number"], po_dt, rec["client"], rec["description"],
            rec["qty_ordered"], rec["qty_received"], rec["uom"],
            rec["unit_price"], rec["total_cost"], rec["invoiced_amount"],
            rec["source_file"]
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws4.cell(row=i, column=c, value=val)
            cell.border = thin_border
            if c in (6,7): cell.number_format = num_fmt_int
            elif c in (9,10,11): cell.number_format = num_fmt_currency
    auto_width(ws4)

    # === TAB 5: Contract Audit ===
    ws5 = wb.create_sheet("Contract Audit")
    headers5 = ["Month", "Description", "Qty Ordered", "Unit Price",
                "Vendor Cost", "Expected Markup", "Expected Invoiced",
                "Actual Invoiced", "Difference", "Flag"]
    style_header(ws5, headers5)
    row_num = 2
    for rec in sorted((r for r in purchase_records if r["month_key"] >= START_MONTH_KEY and normalize_envelope_type(r["description"]) is not None), key=lambda x: x["month_key"]):
        mk = rec["month_key"]
        yr = int(mk.split('-')[0])
        qty = rec["qty_ordered"]
        unit_p = rec["unit_price"]
        vendor_cost = unit_p * (qty / 1000) if qty > 0 else 0
        if yr < 2024:
            expected_markup = vendor_cost * 0.05
        else:
            expected_markup = vendor_cost * 0.122
        expected_invoiced = vendor_cost + expected_markup
        actual_invoiced = rec["invoiced_amount"]
        diff = actual_invoiced - expected_invoiced
        flag = "OK" if abs(diff) < 100 else ("OVER" if diff > 0 else "UNDER")
        row_data = [
            month_key_to_label(mk), rec["description"], qty, unit_p,
            vendor_cost, expected_markup, expected_invoiced,
            actual_invoiced, diff, flag
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws5.cell(row=row_num, column=c, value=val)
            cell.border = thin_border
            if c == 3: cell.number_format = num_fmt_int
            elif c in (4,5,6,7,8,9): cell.number_format = num_fmt_currency
        row_num += 1
    auto_width(ws5)

    # === TAB 6: Data Quality ===
    ws7 = wb.create_sheet("Data Quality")
    headers7 = ["Category", "Source", "Issue"]
    style_header(ws7, headers7)
    for i, (cat, src, issue) in enumerate(data_quality_issues, 2):
        ws7.cell(row=i, column=1, value=cat).border = thin_border
        ws7.cell(row=i, column=2, value=src).border = thin_border
        ws7.cell(row=i, column=3, value=issue).border = thin_border
    auto_width(ws7)

    wb.save(OUTPUT_FILE)
    print(f"Output saved to: {OUTPUT_FILE}")


def main():
    sep = "=" * 80
    print()
    print(sep)
    print("  BROADRIDGE ENVELOPE RECONCILIATION - DATA EXTRACTION")
    print(sep)
    print()

    print("Processing Purchase Reports...")
    process_all_purchase_reports()
    print(f"  Purchase files processed: {files_processed_purchase}")
    print(f"  Purchase files failed: {files_failed_purchase}")
    print(f"  Purchase records extracted: {len(purchase_records):,}")

    # Deduplicate purchase records in two passes:
    # Pass 1: By (date, description, quantity) — catches same-source duplicates
    seen = {}
    unique = []
    dupes = 0
    for rec in purchase_records:
        po_dt = rec["po_date"].strftime("%Y-%m-%d") if rec.get("po_date") else rec["month_key"]
        key = (po_dt, rec["description"], rec["qty_ordered"])
        if key in seen:
            existing_idx = seen[key]
            existing_po = unique[existing_idx].get("po_number", "")
            new_po = rec.get("po_number", "")
            if not existing_po and new_po:
                unique[existing_idx] = rec
            dupes += 1
            continue
        seen[key] = len(unique)
        unique.append(rec)
    # Pass 2: By PO number — catches cross-source duplicates with different descriptions
    po_seen = {}
    final = []
    for rec in unique:
        po = rec.get("po_number", "").strip()
        if po:
            if po in po_seen:
                existing_idx = po_seen[po]
                # Prefer the non-consolidated record (richer data)
                if "2019-2022" in final[existing_idx].get("source_file", ""):
                    final[existing_idx] = rec
                dupes += 1
                continue
            po_seen[po] = len(final)
        final.append(rec)
    if dupes:
        print(f"  Duplicates removed: {dupes}")
        purchase_records.clear()
        purchase_records.extend(final)
    print()

    print("Processing Billing Workbooks...")
    process_all_billing_workbooks()
    print(f"  Billing files processed: {files_processed_billing}")
    print(f"  Billing files failed: {files_failed_billing}")
    print(f"  Volume records extracted: {len(volume_records):,}")
    print(f"  Postage records extracted: {len(postage_records):,}")
    print()

    # YTD file (Apex YTD Envelope Usage) is redundant — its Volume Data and
    # Postage Data tabs duplicate Jan-Aug 2022 billing workbooks exactly,
    # causing double-counting.  Individual billing workbooks cover all months.
    print()

    print()

    apex_records = [r for r in purchase_records if r["month_key"] >= START_MONTH_KEY and normalize_envelope_type(r["description"]) is not None]
    non_apex_qty = sum(r["qty_ordered"] for r in purchase_records if r["month_key"] >= START_MONTH_KEY and normalize_envelope_type(r["description"]) is None)
    if non_apex_qty:
        print(f"  Non-Apex envelopes excluded: {non_apex_qty:,.0f}")
    total_purchased = sum(r["qty_ordered"] for r in apex_records)
    total_received = sum(r["qty_received"] for r in apex_records)
    total_cost = sum(r["total_cost"] for r in apex_records)
    total_invoiced = sum(r["invoiced_amount"] for r in apex_records)
    total_vol_envelopes = sum(r["envelopes"] for r in volume_records if r["month_key"] >= START_MONTH_KEY)
    total_mailed = sum(r["env_mailed"] for r in postage_records if r["month_key"] >= START_MONTH_KEY)
    total_spoils = sum(r["spoils"] for r in postage_records if r["month_key"] >= START_MONTH_KEY)
    variance = total_purchased - total_vol_envelopes

    print(sep)
    print("  SUMMARY")
    print(sep)
    print(f"  Total Envelopes Purchased:      {total_purchased:>14,.0f}")
    print(f"  Total Envelopes Received:       {total_received:>14,.0f}")
    print(f"  Total Envelopes Used (Volume):  {total_vol_envelopes:>14,.0f}")
    print(f"  Total Envelopes Mailed:         {total_mailed:>14,.0f}")
    print(f"  Total Spoils:                   {total_spoils:>14,.0f}")
    print(f"  Net Variance (Purch - Used):    {variance:>14,.0f}")
    if total_purchased > 0:
        print(f"  Variance %:                     {variance/total_purchased:>13.1%}")
    print(f"  Total Purchase Cost:            ${total_cost:>13,.2f}")
    print(f"  Total Invoiced Amount:          ${total_invoiced:>13,.2f}")
    print()

    print()

    if data_quality_issues:
        print(f"  Data Quality Issues: {len(data_quality_issues)}")
        for cat, src, issue in data_quality_issues[:10]:
            print(f"    [{cat}] {src}: {issue}")
        if len(data_quality_issues) > 10:
            print(f"    ... and {len(data_quality_issues)-10} more (see Data Quality tab)")
        print()

    print("Building output Excel workbook...")
    build_output()
    print()
    print(sep)
    print("  COMPLETE")
    print(sep)
    print()

if __name__ == "__main__":
    main()
